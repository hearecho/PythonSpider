#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/7/31 10:39
# @Author  : hearecho
# @Site    : 
# @File    : video_p_aioreq.py
# @Software: PyCharm

'''
    video_p 的异步协程改良版
'''

import requests
import json
import random
from openpyxl import Workbook
import time
import asyncio
import aiohttp

headers = {
    'Host':'api.bilibili.com',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36',

}

query_string_par = {
    'aid':'',
    'jsonp':'jsonp',
}

'''
    对json文件进行了处理    
'''
async def getVideoInfo_plist(url):
    async with aiohttp.ClientSession() as client:
        res = client.get(url,headers=headers)
        # print(json.dumps(r.json(),indent=2))  #美化输出 json
        resu_dict = res.json()
        try:
            datas = resu_dict['data']
            print("总共有{}p".format(len(datas)))
            for data in datas:
                print(data['part'])  # 输出每p的名字
            print("\n")
            return datas
        except:
            print("错误视频av号"+url.split("=")[1])
            return {"status_eror":url.split("=")[1]+"号视频消失"}

async def Savexlsx(datas,ws,i):
    i = i -27960014
    j = 2
    if "status_eror" in datas:
        ws.cell(row = i,column=1,value=datas["status_eror"])
    else:
        ws.cell(row=i,column=1,value=str(len(datas))+"p")
        for data in datas:
            ws.cell(row=i,column=j,value=data['part'])
            j+=1
def Testaio():
    # wb = Workbook()
    # ws = wb.active
    # ws.title = "video_p"
    start = time.time()
    tasks = [asyncio.ensure_future(getVideoInfo_plist('https://api.bilibili.com/x/player/pagelist?aid=' + str(i))) for i in range(27960015,27960100)]

    loop = asyncio.get_event_loop()
    loop.run_until_complete(tasks)


'''
    直接请求得到的数据，json数据
    debug使用
'''
def getVideoInfo_stat(url):
    r = requests.get(url,headers=headers)
    print(json.dumps(r.json(), indent=2))

if __name__ == "__main__":
    wb = Workbook()
    ws = wb.active
    ws.title = "video_p"
    start = time.time()
    for i in range(27960015,27961000):
        plist_url = 'https://api.bilibili.com/x/player/pagelist?aid=' + str(i)
        datas = getVideoInfo_plist(plist_url)
        Savexlsx(datas,ws,i)
        time.sleep(random.randint(0,1))
    # getVideoInfo_stat(plist_url)
    # getVideoInfo_plist(plist_url)
    wb.save("video_p.xlsx")
    end = time.time()

    print("总共耗时{} s".format(end-start))
