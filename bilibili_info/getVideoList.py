#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/8/2 14:38
# @Author  : hearecho
# @Site    : 
# @File    : getVideoList.py
# @Software: PyCharm
'''
    获取b站用户投稿视频列表
    并存取播放数和稿件名字
    url : https://space.bilibili.com/ajax/member/getSubmitVideos?mid=4162287&pagesize=30&tid=0&page=3&keyword=&order=pubdate
    args:{
        mid : 视频作者 UID
        pagesize: 这个有点问题，虽然30 ，50 ，60都可以获取对应数量的稿件，但是再往上会有点问题，所以就使用30就可以了
        page:页码 根据这个构建 url
        tid:稿件属于哪个分区，为0的话 是默认的全部投稿 请求结果里面含有分区号
        后面两个参数不用管
    }

'''
import time
from openpyxl import Workbook
import json
import requests
import matplotlib.pyplot as plt
#headers
headers = {
    "Host":"space.bilibili.com",
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36"
}
#时间转换 从timestamp 到准确的时间
def ChangeTime(timestamp):
    timeArry = time.localtime(timestamp)
    #timeArry : time.struct_time(tm_year=2014, tm_mon=7, tm_mday=16, tm_hour=12, tm_min=9, tm_sec=46, tm_wday=2, tm_yday=197, tm_isdst=0)
    saveStyle = time.strftime("%Y-%m-%d",timeArry)
    return saveStyle

#保存数据，暂时写入到excel文件中，就是一个测试
def SaveFile(context,ws,i):
    # with open("file/1.json","w+",encoding="ascii") as f:
    #     json.dump(Json,f)
    ws.cell(row=i, column=1, value=context["time"])
    ws.cell(row=i, column=2, value=context["play"])
    ws.cell(row=i, column=3, value=context["title"])

#抓取页面 get请求
def getJson(url):
    resp = requests.get(url,headers=headers)
    resp.encoding = resp.apparent_encoding
    status = resp.status_code
    # print(status)
    Json = resp.json()

    return Json["data"]["vlist"]

#绘图
def Plot(x,y):
    plt.bar(range(len(y)),y)
    # plt.ylabel("播放量")
    # plt.title("播放量总表")
    # plt.xticks(range(len(x)),x)
    # plt.ylim([0,9000000])
    plt.show()

if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    ws.title = "video_p"
    x = []
    y = []
    for i in range(1,7):
        url = "https://space.bilibili.com/ajax/member/getSubmitVideos?mid=4162287&pagesize=30&tid=0&page={}&keyword=&order=pubdate".format(i)
        Json = getJson(url)
        for j in range(len(Json)):
            context = {}
            context["title"] = Json[j]["title"]

            context["play"] = Json[j]["play"] if Json[j]["play"] != "--" else 0
            # if context["play"] != "--":
            #     y.append(context["play"])
            # else:
            #     y.append(0)
            # x.append(j+(i-1)*30+1)
            context["time"] = ChangeTime(Json[j]["created"])
            SaveFile(context,ws,j+(i-1)*30+1)
        for i in range(len(Json)):
            print(ChangeTime(Json[i]["created"]),end=" ")
            print(Json[i]["title"],end=" ")
            print(format(Json[i]["play"],">20"))
        print(json.dumps(Json,indent=2,ensure_ascii=False))
    wb.save("file/videoList.xlsx")
    # # print(x)
    # # print(y)
    # Plot(x,y)
