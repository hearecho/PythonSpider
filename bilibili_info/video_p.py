#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/5/26 16:31
# @Author  : hearecho
# @Site    : 
# @File    : video_p.py
# @Software: PyCharm
#视频的 分p信息 以及视频的
import requests
import json
import time
import random
from openpyxl import Workbook
import time
headers = {
    'Host':'api.bilibili.com',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36',

}

query_string_par = {
    'aid':'',
    'jsonp':'jsonp',
}

'''
    对json文件进行了处理    
'''
def getVideoInfo_plist(url):
    r = requests.get(url,headers=headers)
    # print(json.dumps(r.json(),indent=2))  #美化输出 json
    resu_dict = r.json()
    print(resu_dict)
    try:
        datas = resu_dict['data']
        print("总共有{}p".format(len(datas)))
        for data in datas:
            print(data['part'])  # 输出每p的名字
        print("\n")
        return datas
    except:
        print("错误视频av号"+url.split("=")[1])
        return {"status_eror":url.split("=")[1]+"号视频消失"}

'''
    直接请求得到的数据，json数据
'''
def getVideoInfo_stat(url):
    r = requests.get(url,headers=headers)
    print(json.dumps(r.json(), indent=2))
    pass

def Savexlsx(datas,ws,i):
    i = i -27960014
    j = 2
    if "status_eror" in datas:
        ws.cell(row = i,column=1,value=datas["status_eror"])
    else:
        ws.cell(row=i,column=1,value=str(len(datas))+"p")
        for data in datas:
            ws.cell(row=i,column=j,value=data['part'])
            j+=1

if __name__ == "__main__":
    # aid = input("请输入查询的av号:")
    # plist_url = 'https://api.bilibili.com/x/player/pagelist?aid='+aid
    wb = Workbook()
    ws = wb.active
    ws.title = "video_p"
    start = time.time()
    for i in range(27960015,27961017):
        plist_url = 'https://api.bilibili.com/x/player/pagelist?aid=' + str(i)
        datas = getVideoInfo_plist(plist_url)
        Savexlsx(datas,ws,i)
        time.sleep(random.randint(0,1))
    # getVideoInfo_stat(plist_url)
    # getVideoInfo_plist(plist_url)
    wb.save("video_p.xlsx")
    end = time.time()
    # 总共耗时704.2677147388458 s
    print("总共耗时{} s".format(end-start))
